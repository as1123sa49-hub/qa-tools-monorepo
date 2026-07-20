#!/usr/bin/env python3
"""Verify JC platform games against pre-approved gamelist spreadsheet.

Compares work-page XLSX rows with POST /api/fe/v3/game/list response.
Optionally simulates frontend search (case-insensitive substring on game_name).

Usage:
  python tools/verify_jc_games.py --xlsx path/to/gamelist.xlsx
  python tools/verify_jc_games.py --xlsx gamelist.xlsx --env uat --check-search
  python tools/verify_jc_games.py --xlsx gamelist.xlsx --sheet "FA CHAI0626" --provider FC
  python tools/verify_jc_games.py --xlsx gamelist.xlsx --sheet "FA CHAI0626" --stg-only
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

API_URLS = {
    "uat": "https://feg-uat.combosystems.co/api/fe/v3/game/list",
    "pp": "https://feg-pp.combosystems.co/api/fe/v3/game/list",
}

PROVIDER_ALIASES: dict[str, str] = {
    "FA CHAI": "FC",
    "FA CHAI (FC)": "FC",
    "FC": "FC",
    "PRAGMATIC PLAY": "PP",
    "PP": "PP",
    "PG SOFT": "PG",
    "PG": "PG",
    "JILI": "JILI",
    "JDB": "JDB",
    "SEXY CASINO": "SEXY",
    "SEXY": "SEXY",
    "COMBO": "COMBO",
}

CATEGORY_MAP: dict[str, str] = {
    "SLOT": "slot",
    "FISH": "fish",
    "FH": "fish",
    "ARCADE": "arcade",
    "EGAME": "egame",
    "LIVE": "live",
    "TABLE": "table",
}

# Positional fallback when header row cannot be detected (1-based column index)
FALLBACK_COLS = {
    "status": 1,
    "provider": 4,
    "platform": 5,
    "approval": 6,
    "game_code": 7,
    "en_name": 8,
    "zh_name": 9,
    "provider_game_id": 10,
    "category": 12,
    "stg_support": 21,
}

HEADER_HINTS: dict[str, tuple[str, ...]] = {
    "game_code": ("game code", "internal game code", "遊戲代碼", "代碼"),
    "en_name": ("english", "game name", "英文名", "英文"),
    "zh_name": ("chinese", "中文名", "中文"),
    "provider": ("provider", "game provider", "廠商", "content provider"),
    "category": ("stg game type",),
    "approval": ("approved", "approval", "核准"),
    "provider_game_id": ("game id", "table id"),
    "stg_support": ("support for the stg", "測試環境支援度", "測試環境支援"),
}


@dataclass
class SheetGame:
    sheet: str
    row: int
    game_code: str
    en_name: str
    zh_name: str
    provider_label: str
    provider_id: str
    category_label: str
    category: str
    provider_game_id: str
    status: str
    approval: str
    stg_support: str = ""


@dataclass
class ApiGame:
    game_id: str
    game_name: str
    provider_id: str
    category: str
    status: int
    raw: dict


@dataclass
class MatchResult:
    game: SheetGame
    found: bool = False
    api_game: ApiGame | None = None
    issues: list[str] = field(default_factory=list)
    searchable: bool | None = None


def normalize_provider(label: str) -> str:
    text = (label or "").strip().upper()
    if not text:
        return ""
    for key, pid in PROVIDER_ALIASES.items():
        if key in text or text in key:
            return pid
    m = re.search(r"\(([A-Z]+)\)", text)
    if m:
        return m.group(1)
    return text.split()[0] if text else ""


def normalize_category(label: str) -> str:
    text = (label or "").strip().upper()
    return CATEGORY_MAP.get(text, text.lower())


def normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (name or "").lower())


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def detect_header_row(ws, max_scan: int = 20) -> tuple[int, dict[str, int]]:
    max_col = min(ws.max_column or 1, 40)
    for row_idx in range(1, max_scan + 1):
        headers: dict[str, int] = {}
        for col_idx in range(1, max_col + 1):
            raw = cell_str(ws.cell(row_idx, col_idx).value).lower()
            if not raw:
                continue
            for field_name, hints in HEADER_HINTS.items():
                if any(h in raw for h in hints):
                    headers[field_name] = col_idx
        if "game_code" in headers or ("en_name" in headers and "provider" in headers):
            return row_idx, headers
    return 0, {}


def resolve_cols(headers: dict[str, int]) -> dict[str, int]:
    cols = dict(FALLBACK_COLS)
    cols.update(headers)
    return cols


def should_include_sheet(name: str, only: list[str] | None) -> bool:
    if only:
        return any(name == s or s in name for s in only)
    # Skip archived tabs like "1 Pragmatic Play"
    if re.match(r"^\d+\s", name.strip()):
        return False
    return True


def infer_provider_from_sheet(sheet_name: str) -> str:
    base = re.sub(r"\d+$", "", sheet_name).strip()
    return normalize_provider(base)


def is_stg_supported(value: str) -> bool:
    return (value or "").strip().upper() == "Y"


def parse_workbook(
    xlsx_path: Path,
    sheet_filter: list[str] | None = None,
    provider_filter: str | None = None,
    stg_only: bool = False,
) -> list[SheetGame]:
    wb = load_workbook(xlsx_path, data_only=True)
    games: list[SheetGame] = []

    for sheet_name in wb.sheetnames:
        if not should_include_sheet(sheet_name, sheet_filter):
            continue
        ws = wb[sheet_name]
        header_row, headers = detect_header_row(ws)
        cols = resolve_cols(headers)
        data_start = header_row + 1 if header_row else 2
        last_row = ws.max_row or 0

        for row_idx in range(data_start, last_row + 1):
            game_code = cell_str(ws.cell(row_idx, cols["game_code"]).value)
            en_name = cell_str(ws.cell(row_idx, cols["en_name"]).value)
            if not game_code and not en_name:
                continue
            if not re.search(r"[A-Z]{2,}-[A-Z]+-\d+", game_code, re.I) and not en_name:
                continue

            stg_support = cell_str(ws.cell(row_idx, cols["stg_support"]).value)
            if stg_only and not is_stg_supported(stg_support):
                continue

            provider_label = cell_str(ws.cell(row_idx, cols["provider"]).value)
            provider_id = normalize_provider(provider_label) or infer_provider_from_sheet(sheet_name)
            category_label = cell_str(ws.cell(row_idx, cols["category"]).value)
            category = normalize_category(category_label)

            if provider_filter and provider_id != provider_filter.upper():
                continue

            games.append(
                SheetGame(
                    sheet=sheet_name,
                    row=row_idx,
                    game_code=game_code.upper(),
                    en_name=en_name,
                    zh_name=cell_str(ws.cell(row_idx, cols["zh_name"]).value),
                    provider_label=provider_label,
                    provider_id=provider_id,
                    category_label=category_label,
                    category=category,
                    provider_game_id=cell_str(ws.cell(row_idx, cols["provider_game_id"]).value),
                    status=cell_str(ws.cell(row_idx, cols["status"]).value),
                    approval=cell_str(ws.cell(row_idx, cols["approval"]).value),
                    stg_support=stg_support.upper() if stg_support else "",
                )
            )

    wb.close()
    return games


def fetch_api_games(env: str) -> list[ApiGame]:
    url = API_URLS[env]
    req = urllib.request.Request(
        url,
        method="POST",
        data=b"{}",
        headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8"))

    games: list[ApiGame] = []
    for category, items in data.get("d", {}).get("game_type", {}).items():
        if not isinstance(items, list):
            continue
        for item in items:
            games.append(
                ApiGame(
                    game_id=str(item.get("game_id", "")),
                    game_name=item.get("game_name", ""),
                    provider_id=item.get("provider_id", ""),
                    category=category,
                    status=int(item.get("status", 0) or 0),
                    raw=item,
                )
            )
    return games


def build_api_indexes(api_games: list[ApiGame]) -> tuple[dict[str, ApiGame], dict[str, list[ApiGame]]]:
    by_code: dict[str, ApiGame] = {}
    by_name: dict[str, list[ApiGame]] = {}
    for g in api_games:
        code_match = re.search(r"([A-Z]{2,}-[A-Z]+-\d+)", g.game_id, re.I)
        if code_match:
            by_code[code_match.group(1).upper()] = g
        key = normalize_name(g.game_name)
        if key:
            by_name.setdefault(key, []).append(g)
    return by_code, by_name


def find_api_match(
    sheet_game: SheetGame,
    by_code: dict[str, ApiGame],
    by_name: dict[str, list[ApiGame]],
) -> ApiGame | None:
    if sheet_game.game_code and sheet_game.game_code in by_code:
        return by_code[sheet_game.game_code]
    name_key = normalize_name(sheet_game.en_name)
    if name_key and name_key in by_name:
        candidates = [
            g
            for g in by_name[name_key]
            if not sheet_game.provider_id or g.provider_id == sheet_game.provider_id
        ]
        if candidates:
            return candidates[0]
    return None


def simulate_frontend_search(query: str, api_games: list[ApiGame]) -> list[ApiGame]:
    q = (query or "").strip().lower()
    if not q:
        return []
    hits: list[ApiGame] = []
    for g in api_games:
        name = (g.game_name or "").lower()
        if q in name or all(part in name for part in q.split() if len(part) > 1):
            hits.append(g)
    return hits


def verify_games(
    sheet_games: list[SheetGame],
    api_games: list[ApiGame],
    check_search: bool = False,
) -> list[MatchResult]:
    by_code, by_name = build_api_indexes(api_games)
    results: list[MatchResult] = []

    for sg in sheet_games:
        result = MatchResult(game=sg)
        api = find_api_match(sg, by_code, by_name)

        if not api:
            result.issues.append("list API 中找不到此遊戲")
            results.append(result)
            if check_search:
                hits = simulate_frontend_search(sg.en_name, api_games)
                result.searchable = any(
                    h.provider_id == sg.provider_id for h in hits
                ) if sg.provider_id else bool(hits)
                if not result.searchable:
                    result.issues.append("前台搜尋不到（以英文名模擬）")
            continue

        result.found = True
        result.api_game = api

        if sg.provider_id and api.provider_id != sg.provider_id:
            result.issues.append(
                f"廠商不符：試算表={sg.provider_id} API={api.provider_id}"
            )
        if sg.category and api.category != sg.category:
            result.issues.append(
                f"類型不符：試算表={sg.category} API={api.category}"
            )
        if sg.en_name and normalize_name(sg.en_name) != normalize_name(api.game_name):
            result.issues.append(
                f"英文名不符：試算表='{sg.en_name}' API='{api.game_name}'"
            )
        if api.status != 1:
            result.issues.append(f"API status={api.status}（非上架狀態）")

        if check_search:
            hits = simulate_frontend_search(sg.en_name, api_games)
            result.searchable = any(h.game_id == api.game_id for h in hits)
            if not result.searchable:
                result.issues.append("前台搜尋不到（以英文名模擬）")

        results.append(result)

    return results


def print_report(results: list[MatchResult], check_search: bool, stg_only: bool) -> int:
    total = len(results)
    passed = sum(1 for r in results if r.found and not r.issues)
    missing = sum(1 for r in results if not r.found)
    with_issues = sum(1 for r in results if r.found and r.issues)

    print(f"\n{'=' * 60}")
    print(f"驗證結果：{passed}/{total} 通過")
    print(f"  找不到：{missing}  有差異：{with_issues}")
    if check_search:
        searchable = sum(1 for r in results if r.searchable)
        print(f"  可搜尋：{searchable}/{total}")
    if stg_only:
        print("  篩選：僅 STG 支援度 = Y")
    print(f"{'=' * 60}")

    by_sheet: dict[str, list[MatchResult]] = {}
    for r in results:
        by_sheet.setdefault(r.game.sheet, []).append(r)

    for sheet, items in sorted(by_sheet.items()):
        sheet_pass = sum(1 for r in items if r.found and not r.issues)
        print(f"\n## {sheet} ({sheet_pass}/{len(items)})")
        for r in items:
            g = r.game
            if r.found and not r.issues:
                api = r.api_game
                search_tag = ""
                if check_search:
                    search_tag = " [search:OK]" if r.searchable else " [search:FAIL]"
                stg_tag = f" [STG:{g.stg_support or '-'}]" if g.stg_support or stg_only else ""
                print(
                    f"  [OK] {g.game_code} | {g.en_name} -> "
                    f"{api.game_id} ({api.game_name}){stg_tag}{search_tag}"
                )
            else:
                status = "[FAIL]" if not r.found else "[WARN]"
                stg_tag = f" STG={g.stg_support or '-'}" if g.stg_support or stg_only else ""
                print(
                    f"  {status} [{g.row}] {g.game_code} | {g.en_name} "
                    f"({g.provider_id}/{g.category}){stg_tag}"
                )
                for issue in r.issues:
                    print(f"      - {issue}")
                if r.api_game:
                    print(f"      API: {r.api_game.game_id} ({r.api_game.game_name})")

    return 0 if passed == total else 1


def export_json(results: list[MatchResult], path: Path) -> None:
    payload = []
    for r in results:
        item = {
            "sheet": r.game.sheet,
            "row": r.game.row,
            "game_code": r.game.game_code,
            "en_name": r.game.en_name,
            "zh_name": r.game.zh_name,
            "provider_id": r.game.provider_id,
            "category": r.game.category,
            "stg_support": r.game.stg_support,
            "found": r.found,
            "searchable": r.searchable,
            "issues": r.issues,
        }
        if r.api_game:
            item["api"] = {
                "game_id": r.api_game.game_id,
                "game_name": r.api_game.game_name,
                "provider_id": r.api_game.provider_id,
                "category": r.api_game.category,
                "status": r.api_game.status,
            }
        payload.append(item)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nJSON 報告已寫入：{path}")


def _configure_stdout() -> None:
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8")
            except Exception:
                pass


def main() -> int:
    _configure_stdout()
    parser = argparse.ArgumentParser(
        description="Verify JC platform games against pre-approved gamelist XLSX"
    )
    parser.add_argument("--xlsx", required=True, help="Path to pre-approved gamelist .xlsx")
    parser.add_argument("--env", choices=["uat", "pp"], default="uat", help="API environment")
    parser.add_argument("--sheet", action="append", help="Only verify matching sheet name(s)")
    parser.add_argument("--provider", help="Only verify this provider_id (e.g. FC, PP)")
    parser.add_argument(
        "--check-search",
        action="store_true",
        help="Simulate frontend search by English game name",
    )
    parser.add_argument(
        "--stg-only",
        action="store_true",
        help="Only verify rows where STG support column (U) = Y",
    )
    parser.add_argument("--output-json", help="Write detailed results to JSON file")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found", file=sys.stderr)
        return 1

    print(f"讀取試算表：{xlsx_path}")
    if args.stg_only:
        print("  篩選：僅驗證 STG 支援度 = Y 的遊戲")
    sheet_games = parse_workbook(
        xlsx_path, args.sheet, args.provider, stg_only=args.stg_only
    )
    print(f"  試算表遊戲數：{len(sheet_games)}")
    if not sheet_games:
        print("ERROR: 未解析到任何遊戲列，請確認工作表格式", file=sys.stderr)
        return 1

    by_provider: dict[str, int] = {}
    for g in sheet_games:
        by_provider[g.provider_id] = by_provider.get(g.provider_id, 0) + 1
    for pid, count in sorted(by_provider.items()):
        print(f"    {pid}: {count}")

    print(f"\n呼叫 list API ({args.env})...")
    api_games = fetch_api_games(args.env)
    print(f"  API 遊戲總數：{len(api_games)}")

    results = verify_games(sheet_games, api_games, check_search=args.check_search)
    exit_code = print_report(results, args.check_search, args.stg_only)

    if args.output_json:
        export_json(results, Path(args.output_json))

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
